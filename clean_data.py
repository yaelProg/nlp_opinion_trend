from __future__ import annotations

from openpyxl import Workbook, load_workbook

import config
from text_utils import clean_text


def word_count(text: str) -> int:
    t = text.strip()
    if not t:
        return 0
    return len(t.split())


def clean_excel(input_path: str, output_path: str, min_words: int = 4) -> tuple[int, int]:
    wb_in = load_workbook(input_path)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "clean_data"

    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        ws_out.append(["id", "type", "parent_id", "text", "city", "timestamp"])
        wb_out.save(output_path)
        return 0, 0

    header = list(rows[0])
    ws_out.append(header)

    try:
        text_idx = header.index("text")
    except ValueError as e:
        raise RuntimeError("Input file must include a 'text' column.") from e

    kept = 0
    removed = 0
    for row in rows[1:]:
        values = list(row)
        while len(values) < len(header):
            values.append("")

        text = str(values[text_idx] or "")
        cleaned = clean_text(text)
        if word_count(cleaned) < min_words:
            removed += 1
            continue

        values[text_idx] = cleaned
        ws_out.append(values)
        kept += 1

    wb_out.save(output_path)
    return kept, removed


def main() -> None:
    kept, removed = clean_excel(config.INPUT_FILE, config.CLEAN_OUTPUT_FILE, min_words=config.MIN_WORDS)
    print(f"Saved {kept} cleaned rows to {config.CLEAN_OUTPUT_FILE} (removed {removed} rows)")


if __name__ == "__main__":
    main()
